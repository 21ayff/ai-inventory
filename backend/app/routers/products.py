"""商品管理接口"""
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import or_
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook

from ..database import get_db
from ..models import Category, Product
from ..schemas import ProductCreate, ProductUpdate, ProductOut
from ..dependencies import get_current_user
from ..stock_math import calc_stock_params, DEFAULT_TARGET_DAYS, DEFAULT_Z_SCORE
from ..sales_stats import calc_sales_std
from ..eoq import calc_eoq, get_order_cost, get_holding_cost_rate

router = APIRouter(prefix="/api/products", tags=["products"], dependencies=[Depends(get_current_user)])


def _category_strategy(db: Session, category_id: int | None):
    """读取分类的库存策略，未选择分类时返回默认值"""
    if category_id:
        cat = db.query(Category).filter(Category.id == category_id).first()
        if cat:
            return cat.target_days, cat.z_score
    return DEFAULT_TARGET_DAYS, DEFAULT_Z_SCORE


@router.get("", response_model=list[ProductOut])
def list_products(
    search: str = "",
    stock_filter: str = "",
    db: Session = Depends(get_db),
):
    """商品列表，支持搜索和库存筛选"""
    query = db.query(Product).filter(Product.deleted == False)

    if search:
        query = query.filter(
            or_(Product.name.contains(search), Product.sku.contains(search))
        )

    if stock_filter == "low":      # 库存不足
        query = query.filter(Product.current_stock < Product.min_stock)

    return query.order_by(Product.id.desc()).all()


@router.post("", response_model=ProductOut)
def create_product(data: ProductCreate, db: Session = Depends(get_db)):
    """新增商品（安全库存/订货点/建议补货量由 AI 自动计算）"""
    if not data.name:
        raise HTTPException(status_code=400, detail="商品名称不能为空")

    # 空字符串 SKU 视为未填写，转为 None 避免触发 unique 冲突
    if data.sku is not None:
        data.sku = data.sku.strip()
        if data.sku == "":
            data.sku = None

    if data.sku:
        exists = db.query(Product).filter(
            Product.sku == data.sku, Product.deleted == False
        ).first()
        if exists:
            raise HTTPException(status_code=400, detail="SKU 已存在")

    # 读取分类策略，自动计算三个专业参数
    target_days, z_score = _category_strategy(db, data.category_id)
    # 读取全局订货成本和持有成本率，计算经济订货量 EOQ
    order_cost = get_order_cost(db)
    holding_cost_rate = get_holding_cost_rate(db)
    eoq_value = calc_eoq(data.daily_sales, data.cost_price, order_cost, holding_cost_rate)
    # 新增商品时还没有历史出库记录，sales_std 传 None，自动回退到估算（日均销量 × 20%）
    min_stock, rop, eoq = calc_stock_params(
        daily_sales=data.daily_sales,
        lead_time_days=data.lead_time_days,
        current_stock=data.current_stock,
        category_target_days=target_days,
        category_z_score=z_score,
        shelf_life_days=data.shelf_life_days,
        sales_std=None,
        eoq=eoq_value,
    )

    product = Product(
        name=data.name,
        sku=data.sku,
        category_id=data.category_id,
        unit=data.unit,
        current_stock=data.current_stock,
        daily_sales=data.daily_sales,
        lead_time_days=data.lead_time_days,
        shelf_life_days=data.shelf_life_days,
        cost_price=data.cost_price,
        supplier_name=data.supplier_name,
        supplier_phone=data.supplier_phone,
        min_stock=min_stock,
        rop=rop,
        eoq=eoq,
    )
    db.add(product)
    db.commit()
    db.refresh(product)
    return product


@router.get("/import-template")
def download_import_template():
    """下载 Excel 导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "商品导入模板"
    ws.append(["商品名称", "SKU", "单位", "当前库存", "日均销量", "平均到货天数", "保质期天数", "成本价"])
    ws.append(["示例商品", "SKU001", "个", 100, 10, 7, 90, 5.0])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=import_template.xlsx"},
    )


@router.post("/import")
async def import_products(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """从 Excel 批量导入商品，按名称覆盖更新已有商品"""
    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="无法读取 Excel 文件，请使用 .xlsx 格式")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头

    def parse_number(v):
        """解析数值：空值返回 0，非法值返回 None"""
        if v is None:
            return 0.0
        s = str(v).strip()
        if s == "":
            return 0.0
        try:
            return float(s)
        except (ValueError, TypeError):
            return None

    def parse_optional_number(v):
        """解析可空数值：空值返回 None，非法值返回 None"""
        if v is None:
            return None
        s = str(v).strip()
        if s == "":
            return None
        try:
            return float(s)
        except (ValueError, TypeError):
            return None

    success = 0
    failed = []

    for i, row in enumerate(rows, start=2):
        # 跳过完全空的行
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        name = str(row[0]).strip() if len(row) > 0 and row[0] else ""
        sku = str(row[1]).strip() if len(row) > 1 and row[1] else None
        unit = str(row[2]).strip() if len(row) > 2 and row[2] else "个"
        current_stock = parse_number(row[3] if len(row) > 3 else None)
        daily_sales = parse_number(row[4] if len(row) > 4 else None)
        lead_time_days = parse_number(row[5] if len(row) > 5 else None)
        shelf_life_days = parse_optional_number(row[6] if len(row) > 6 else None)
        cost_price = parse_optional_number(row[7] if len(row) > 7 else None)

        if not name:
            failed.append(f"第{i}行：商品名称为空")
            continue

        if any(v is None for v in [current_stock, daily_sales, lead_time_days]):
            failed.append(f"第{i}行：数值格式错误")
            continue

        # 导入时不含分类信息，使用默认策略计算三个专业参数
        target_days, z_score = _category_strategy(db, None)
        # 读取全局订货成本和持有成本率，计算经济订货量 EOQ
        order_cost = get_order_cost(db)
        holding_cost_rate = get_holding_cost_rate(db)
        eoq_value = calc_eoq(daily_sales, cost_price, order_cost, holding_cost_rate)
        min_stock, rop, eoq = calc_stock_params(
            daily_sales=daily_sales,
            lead_time_days=lead_time_days,
            current_stock=current_stock,
            category_target_days=target_days,
            category_z_score=z_score,
            shelf_life_days=shelf_life_days,
            eoq=eoq_value,
        )

        # 按名称查找已有商品
        existing = db.query(Product).filter(
            Product.name == name, Product.deleted == False
        ).first()

        # SKU 唯一性检查（排除当前正在更新的商品）
        if sku:
            sku_exists = db.query(Product).filter(
                Product.sku == sku,
                Product.deleted == False,
                Product.id != (existing.id if existing else -1),
            ).first()
            if sku_exists:
                failed.append(f"第{i}行：SKU {sku} 已被其他商品使用")
                continue

        if existing:
            # 覆盖更新已有商品
            existing.sku = sku
            existing.unit = unit
            existing.current_stock = current_stock
            existing.daily_sales = daily_sales
            existing.lead_time_days = lead_time_days
            existing.shelf_life_days = shelf_life_days
            existing.cost_price = cost_price
            existing.min_stock = min_stock
            existing.rop = rop
            existing.eoq = eoq
        else:
            # 新增商品
            db.add(
                Product(
                    name=name,
                    sku=sku,
                    unit=unit,
                    current_stock=current_stock,
                    daily_sales=daily_sales,
                    lead_time_days=lead_time_days,
                    shelf_life_days=shelf_life_days,
                    cost_price=cost_price,
                    min_stock=min_stock,
                    rop=rop,
                    eoq=eoq,
                )
            )

        success += 1

    db.commit()

    return {
        "success": success,
        "failed_count": len(failed),
        "failed": failed,
    }


@router.get("/{product_id}", response_model=ProductOut)
def get_product(product_id: int, db: Session = Depends(get_db)):
    """商品详情"""
    product = db.query(Product).filter(
        Product.id == product_id, Product.deleted == False
    ).first()
    if not product:
        raise HTTPException(status_code=404, detail="商品不存在")
    return product


@router.put("/{product_id}", response_model=ProductOut)
def update_product(product_id: int, data: ProductUpdate, db: Session = Depends(get_db)):
    """编辑商品（安全库存/订货点/建议补货量由 AI 自动计算）"""
    product = db.query(Product).filter(
        Product.id == product_id, Product.deleted == False
    ).first()
    if not product:
        raise HTTPException(status_code=404, detail="商品不存在")

    update_data = data.model_dump(exclude_unset=True)

    # 空字符串 SKU 视为未填写，转为 None 避免触发 unique 冲突
    if "sku" in update_data and update_data["sku"] is not None:
        update_data["sku"] = update_data["sku"].strip()
        if update_data["sku"] == "":
            update_data["sku"] = None

    if update_data.get("sku"):
        exists = db.query(Product).filter(
            Product.sku == update_data["sku"],
            Product.id != product_id,
            Product.deleted == False,
        ).first()
        if exists:
            raise HTTPException(status_code=400, detail="SKU 已存在")

    for key, value in update_data.items():
        setattr(product, key, value)

    # 读取分类策略，重新计算库存参数（日均销量为 0 时三个参数都归 0）
    target_days, z_score = _category_strategy(db, product.category_id)
    # 读取全局订货成本和持有成本率，计算经济订货量 EOQ
    order_cost = get_order_cost(db)
    holding_cost_rate = get_holding_cost_rate(db)
    eoq_value = calc_eoq(product.daily_sales, product.cost_price, order_cost, holding_cost_rate)
    # 查询历史销量标准差（数据不足时返回 None，自动回退到估算）
    sales_std, _, _ = calc_sales_std(db, product.id)
    product.min_stock, product.rop, product.eoq = calc_stock_params(
        daily_sales=product.daily_sales,
        lead_time_days=product.lead_time_days,
        current_stock=product.current_stock,
        category_target_days=target_days,
        category_z_score=z_score,
        shelf_life_days=product.shelf_life_days,
        sales_std=sales_std,
        eoq=eoq_value,
    )

    db.commit()
    db.refresh(product)
    return product


@router.delete("/{product_id}")
def delete_product(product_id: int, db: Session = Depends(get_db)):
    """删除商品（软删除）"""
    product = db.query(Product).filter(
        Product.id == product_id, Product.deleted == False
    ).first()
    if not product:
        raise HTTPException(status_code=404, detail="商品不存在")

    product.deleted = True
    # 释放 SKU，允许以后重建相同 SKU 的商品
    product.sku = None
    db.commit()
    return {"message": "删除成功"}
